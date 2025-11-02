# validador_fiscal/tools/report_generator.py
"""
Gerador de Relatório Excel com Divergências
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
import os
from datetime import datetime


def gerar_relatorio_excel(relatorio, output_path=None):
    """Gera Excel com divergências destacadas"""
    
    if not output_path:
        os.makedirs("data/reports/excel", exist_ok=True)
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        nf_numero = relatorio.get("metadata", {}).get("numero", "SN")
        output_path = f"data/reports/excel/relatorio_NF{nf_numero}_{timestamp}.xlsx"
    
    print(f"📊 Gerando Excel: {output_path}")
    
    with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
        # ABA 1: Resumo
        _criar_resumo(writer, relatorio)
        # ABA 2: Divergências
        _criar_divergencias(writer, relatorio)
        # ABA 3: Totais
        _criar_totais(writer, relatorio)
    
    _formatar_excel(output_path)
    
    print(f"✅ Excel: {output_path} ({os.path.getsize(output_path)/1024:.1f} KB)")
    return output_path


def _criar_resumo(writer, relatorio):
    metadata = relatorio.get("metadata", {})
    resumo = relatorio.get("resumo_executivo", {})
    
    dados = {
        "Item": ["Chave", "Número", "Total Itens", "Total Calculado", "Total Declarado", "Divergência", "Risco"],
        "Valor": [
            metadata.get("chave", "")[:50],
            metadata.get("numero", ""),
            resumo.get("total_itens", 0),
            f"R$ {resumo.get('total_calculado', 0):,.2f}",
            f"R$ {resumo.get('total_declarado', 0):,.2f}",
            f"R$ {resumo.get('divergencia_absoluta', 0):,.2f}",
            resumo.get("nivel_risco", "")
        ]
    }
    
    pd.DataFrame(dados).to_excel(writer, sheet_name='Resumo', index=False)


def _criar_divergencias(writer, relatorio):
    itens = relatorio.get("itens", [])[:10000]  # Max 10k
    
    diverg = []
    for item in itens:
        impostos = item.get("impostos", {})
        for imp, dados in impostos.items():
            val = dados.get("valor", 0)
            if val > 10:  # Só se > R$ 10
                diverg.append({
                    "Item": item.get("numero"),
                    "Código": item.get("codigo"),
                    "Imposto": imp,
                    "Valor": val
                })
    
    if not diverg:
        diverg = [{"Item": "", "Código": "", "Imposto": "✅ SEM DIVERGÊNCIAS", "Valor": 0}]
    
    pd.DataFrame(diverg).to_excel(writer, sheet_name='Divergências', index=False)


def _criar_totais(writer, relatorio):
    totais = relatorio.get("totais_por_imposto", {})
    
    dados = [
        {"Imposto": imp, "Calculado": v.get("calculado", 0), "Declarado": v.get("declarado", 0)}
        for imp, v in totais.items()
    ]
    
    pd.DataFrame(dados).to_excel(writer, sheet_name='Totais', index=False)


def _formatar_excel(path):
    wb = load_workbook(path)
    
    for ws in wb.worksheets:
        # Header azul
        for cell in ws[1]:
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Ajustar colunas
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)
    
    wb.save(path)